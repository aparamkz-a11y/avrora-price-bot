import io
import math
import re
from pathlib import Path
from typing import Optional

import openpyxl

from suppliers_config import AVRORA

# ── Markup formulas ───────────────────────────────────────────────────────────

def _retail_pct(price: float) -> float:
    if price < 5_000:      return 0.30
    if price < 10_000:     return 0.25
    if price < 18_000:     return 0.18
    if price < 40_000:     return 0.15
    if price < 70_000:     return 0.11
    if price < 120_000:    return 0.08
    if price < 250_000:    return 0.06
    if price < 500_000:    return 0.05
    if price < 1_000_000:  return 0.035
    return 0.025


def _round_step(value: float) -> int:
    if value < 5_000:      step = 10
    elif value < 50_000:   step = 50
    elif value < 500_000:  step = 100
    else:                  step = 1_000
    return int(math.ceil(value / step) * step)


def _apply_retail(base: float) -> Optional[int]:
    if not base or base <= 0:
        return None
    markup = max(base * _retail_pct(base), 100)
    return _round_step(base + markup)


def _apply_wholesale(base: float) -> Optional[int]:
    if not base or base <= 0:
        return None
    markup = max(base * _retail_pct(base) * 0.60, 50)
    return _round_step(base + markup)


# ── Price column detection ────────────────────────────────────────────────────

PRICE_KEYWORDS = [
    "цена", "прайс", "стоимость", "price", "тенге",
    "тг/м", "тг/шт", "тг/кг", "тг/т", "тг.",
    "руб", "cost", "итого", "сумма",
]

# Values likely to be prices in metal trade (tenge per unit)
_PRICE_MIN = 200
_PRICE_MAX = 15_000_000


def _find_price_cols(ws) -> list[int]:
    """
    Find columns that contain prices.
    Strategy 1: look for price keywords in any of the first 6 rows.
    Strategy 2: if nothing found, use value distribution heuristic.
    """
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # --- Strategy 1: keyword scan in header rows ---
    found_by_keyword: list[int] = []
    for row in ws.iter_rows(max_row=min(6, max_row), values_only=True):
        for col_idx, val in enumerate(row):
            if val and isinstance(val, str):
                v = val.lower().strip()
                if any(kw in v for kw in PRICE_KEYWORDS):
                    found_by_keyword.append(col_idx)
    if found_by_keyword:
        return list(dict.fromkeys(found_by_keyword))  # deduplicated, order preserved

    # --- Strategy 2: value-based heuristic ---
    # For each column count how many cells look like prices
    col_price_count = [0] * max_col
    col_total_count = [0] * max_col
    for row in ws.iter_rows(min_row=2, values_only=True):
        for col_idx, val in enumerate(row):
            if col_idx >= max_col:
                break
            if isinstance(val, (int, float)) and val > 0:
                col_total_count[col_idx] += 1
                if _PRICE_MIN <= val <= _PRICE_MAX:
                    col_price_count[col_idx] += 1

    # Column is a price column if ≥60% of its numeric values look like prices
    # and it has at least 3 numeric values, and it's not one of the first 2 cols
    price_cols = []
    for col_idx in range(min(2, max_col), max_col):
        total = col_total_count[col_idx]
        if total >= 3 and col_price_count[col_idx] / total >= 0.60:
            price_cols.append(col_idx)

    # If too many columns matched, take only those with highest match ratio
    if len(price_cols) > 3:
        price_cols.sort(key=lambda c: col_price_count[c] / max(col_total_count[c], 1), reverse=True)
        price_cols = price_cols[:3]

    return sorted(price_cols)


# ── Contact replacement ───────────────────────────────────────────────────────

_PHONE_RE = re.compile(r'(\+7|8)[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}')
_EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')


def _replace_contacts(ws):
    line1 = (
        f"{AVRORA['company']}  |  "
        f"Тел: {AVRORA['phone']}  |  "
        f"WhatsApp: {AVRORA['whatsapp']}"
    )
    line2 = (
        f"{AVRORA['email']}  |  "
        f"{AVRORA['website']}  |  "
        f"{AVRORA['address']}  |  "
        f"{AVRORA['schedule']}"
    )
    replaced = 0
    for row in ws.iter_rows(max_row=6):
        for cell in row:
            if not cell.value:
                continue
            val = str(cell.value)
            if _PHONE_RE.search(val) or _EMAIL_RE.search(val) or "www." in val.lower():
                cell.value = line1 if replaced == 0 else line2
                replaced += 1
                if replaced >= 2:
                    return
    if replaced == 0:
        ws.cell(row=1, column=1).value = line1


# ── Main processing function ──────────────────────────────────────────────────

def process_file(filepath: str) -> tuple[bytes, bytes, int]:
    """
    Process any supplier xlsx file.
    Returns (retail_bytes, wholesale_bytes, price_cols_count).
    Works with ANY Excel price list — no supplier config required.
    """
    results = []
    price_cols_count = 0

    for markup_fn in (_apply_retail, _apply_wholesale):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        price_cols = _find_price_cols(ws)
        price_cols_count = len(price_cols)

        for row in ws.iter_rows():
            for col_idx, cell in enumerate(row):
                if col_idx not in price_cols:
                    continue
                if not isinstance(cell.value, (int, float)):
                    continue
                base = float(cell.value)
                if base < _PRICE_MIN or base > _PRICE_MAX:
                    continue
                new_price = markup_fn(base)
                if new_price is not None:
                    cell.value = new_price

        _replace_contacts(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        results.append(buf.read())

    return results[0], results[1], price_cols_count
